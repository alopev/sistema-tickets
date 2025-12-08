from flask import Blueprint, render_template, request, redirect, url_for, send_file
from flask_login import login_required, current_user
from app import db
from app.models import Ticket, User
from app.utils import tech_required
from app.utils.alerts import success, error, warning
import pandas as pd  # Commented temporarily - install pandas later for Excel/CSV export
from fpdf import FPDF
import io
from datetime import datetime

bp = Blueprint('main', __name__)

@bp.route('/')
@login_required
def index():
    # Base query
    query = Ticket.query
    
    # Filter based on role
    if current_user.role == 'tecnico':
        query = query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        query = query.filter_by(created_by_id=current_user.id)
        
    # Calculate stats
    total_tickets = query.count()
    open_tickets = query.filter_by(status='abierto').count()
    in_progress = query.filter_by(status='en_proceso').count()
    closed_tickets = query.filter_by(status='cerrado').count()
    
    # Data for charts
    status_data = [open_tickets, in_progress, closed_tickets]
    
    # Priority Data (re-apply filters for these counts)
    # Note: We can't reuse the 'query' object easily for different filters if it's already filtered by status above
    # So we rebuild the base filter
    base_query = Ticket.query
    if current_user.role == 'tecnico':
        base_query = base_query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        base_query = base_query.filter_by(created_by_id=current_user.id)

    high_priority = base_query.filter_by(priority='alta').count()
    medium_priority = base_query.filter_by(priority='media').count()
    low_priority = base_query.filter_by(priority='baja').count()
    priority_data = [high_priority, medium_priority, low_priority]
    
    return render_template('dashboard.html', 
                         total=total_tickets, 
                         open=open_tickets, 
                         process=in_progress, 
                         closed=closed_tickets,
                         status_data=status_data,
                         priority_data=priority_data)

@bp.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    # Re-calculate stats (logic duplicated for now, could be refactored)
    query = Ticket.query
    if current_user.role == 'tecnico':
        query = query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        query = query.filter_by(created_by_id=current_user.id)
        
    total = query.count()
    open_t = query.filter_by(status='abierto').count()
    process = query.filter_by(status='en_proceso').count()
    closed = query.filter_by(status='cerrado').count()
    
    # Priority
    base_query = Ticket.query
    if current_user.role == 'tecnico':
        base_query = base_query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        base_query = base_query.filter_by(created_by_id=current_user.id)

    high = base_query.filter_by(priority='alta').count()
    medium = base_query.filter_by(priority='media').count()
    low = base_query.filter_by(priority='baja').count()
    
    return {
        'total': total,
        'open': open_t,
        'process': process,
        'closed': closed,
        'status_data': [open_t, process, closed],
        'priority_data': [high, medium, low]
    }

@bp.route('/tickets')
@login_required
def tickets():
    if current_user.role == 'usuario':
        return redirect(url_for('main.my_tickets'))
    
    status_filter = request.args.get('status')
    if status_filter:
        tickets = Ticket.query.filter_by(status=status_filter).all()
    else:
        tickets = Ticket.query.all()
        
    return render_template('tickets/list.html', tickets=tickets, title='Todos los Tickets')

@bp.route('/my_tickets')
@login_required
def my_tickets():
    tickets = Ticket.query.filter_by(created_by_id=current_user.id).all()
    return render_template('tickets/list.html', tickets=tickets, title='Mis Tickets')

@bp.route('/ticket/create', methods=['GET', 'POST'])
@login_required
def create_ticket():
    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        priority = request.form['priority']
        
        file = request.files.get('file')
        filename = None
        if file and file.filename:
            from werkzeug.utils import secure_filename
            from app.security import validate_file_extension
            
            # Validate file extension
            allowed_extensions = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'txt', 'zip'}
            if not validate_file_extension(file.filename, allowed_extensions):
                error('Tipo de archivo no permitido. Extensiones permitidas: PDF, PNG, JPG, GIF, DOC, DOCX, TXT, ZIP')
                return redirect(url_for('main.create_ticket'))
            
            filename = secure_filename(file.filename)
            # Ensure upload directory exists
            import os
            from flask import current_app
            os.makedirs(current_app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
        
        # Generate ticket number
        ticket_number = Ticket.generate_ticket_number()
        
        ticket = Ticket(
            ticket_number=ticket_number,
            title=title, 
            description=description, 
            priority=priority, 
            created_by_id=current_user.id, 
            attachment=filename
        )
        db.session.add(ticket)
        db.session.commit()
        
        # Emit dashboard update
        from app import socketio
        socketio.emit('dashboard_update')
        
        success(f'Ticket {ticket_number} creado exitosamente')
        return redirect(url_for('main.my_tickets'))
        
    return render_template('tickets/create.html')

@bp.route('/ticket/<int:id>', methods=['GET', 'POST'])
@login_required
def ticket_detail(id):
    ticket = Ticket.query.get_or_404(id)
    technicians = User.query.filter_by(role='tecnico').all()
    
    if request.method == 'POST':
        if current_user.role in ['admin', 'tecnico']:
            status = request.form.get('status')
            assigned_to_id = request.form.get('assigned_to')
            
            # Track if assignment changed for email notification
            assignment_changed = False
            newly_assigned_user = None
            
            if status:
                ticket.status = status
            if assigned_to_id:
                old_assigned_id = ticket.assigned_to_id
                ticket.assigned_to_id = int(assigned_to_id) if assigned_to_id else None
                
                # Check if assignment actually changed
                if old_assigned_id != ticket.assigned_to_id and ticket.assigned_to_id:
                    assignment_changed = True
                    newly_assigned_user = User.query.get(ticket.assigned_to_id)
                
            db.session.commit()
            
            # Send email notification if ticket was assigned
            if assignment_changed and newly_assigned_user:
                from app.email import send_ticket_assigned_email
                send_ticket_assigned_email(ticket, newly_assigned_user)
            
            # Emit update event
            from app import socketio
            
            # Emit dashboard update
            socketio.emit('dashboard_update')
            
            success('Ticket actualizado')
            return redirect(url_for('main.ticket_detail', id=ticket.id))
            
    return render_template('tickets/detail.html', ticket=ticket, technicians=technicians)

@bp.route('/tickets/<int:id>/details')
@login_required
def ticket_details_json(id):
    """Return ticket details as JSON for AJAX loading in modal"""
    ticket = Ticket.query.get_or_404(id)
    
    # Map status to Spanish
    status_labels = {
        'abierto': 'Abierto',
        'en_proceso': 'En Proceso',
        'cerrado': 'Cerrado'
    }
    
    return {
        'ticket_number': ticket.ticket_number,
        'title': ticket.title,
        'description': ticket.description,
        'status': ticket.status,
        'status_label': status_labels.get(ticket.status, ticket.status),
        'priority': ticket.priority.capitalize() if ticket.priority else 'Media',
        'created_by': ticket.created_by.username,
        'created_at': ticket.created_at.strftime('%d/%m/%Y %H:%M'),
        'assigned_to': ticket.assigned_to.username if ticket.assigned_to else None,
        'assigned_to_id': ticket.assigned_to_id
    }

@bp.route('/api/technicians')
@login_required
def get_technicians():
    """Return list of technicians for dropdown in edit modal"""
    technicians = User.query.filter_by(role='tecnico').all()
    return [{'id': t.id, 'username': t.username, 'role': t.role} for t in technicians]

@bp.route('/ticket/<int:id>/comment', methods=['POST'])
@login_required
def add_comment(id):
    content = request.form.get('content')
    if content:
        from app.models import Comment
        from app.security import sanitize_html
        
        # Sanitize HTML content to prevent XSS
        sanitized_content = sanitize_html(content)
        
        ticket = Ticket.query.get_or_404(id)
        comment = Comment(content=sanitized_content, user_id=current_user.id, ticket_id=id)
        db.session.add(comment)
        db.session.commit()
        
        # Send email notifications to relevant users
        from app.email import send_ticket_comment_email
        recipients = set()  # Use set to avoid duplicates
        
        # Notify ticket creator if they didn't write the comment
        if ticket.created_by_id != current_user.id:
            recipients.add(ticket.created_by)
        
        # Notify assigned technician if they didn't write the comment
        if ticket.assigned_to and ticket.assigned_to_id != current_user.id:
            recipients.add(ticket.assigned_to)
        
        # Send emails
        for recipient in recipients:
            send_ticket_comment_email(ticket, current_user, sanitized_content, recipient)
        
        # Emit new comment event
        from app import socketio
        socketio.emit('new_comment', {
            'ticket_id': id,
            'user': current_user.username,
            'content': sanitized_content,
            'timestamp': comment.created_at.strftime('%Y-%m-%d %H:%M')
        })
        
        success('Comentario agregado')
    return redirect(url_for('main.ticket_detail', id=id))

@bp.route('/export/excel')
@login_required
def export_excel():
    # Filter based on role
    query = Ticket.query
    if current_user.role == 'tecnico':
        query = query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        query = query.filter_by(created_by_id=current_user.id)
        
    tickets = query.all()
    data = [{
        'ID': t.id,
        'Título': t.title,
        'Estado': t.status,
        'Prioridad': t.priority,
        'Creado Por': t.created_by.username,
        'Asignado A': t.assigned_to.username if t.assigned_to else 'Sin asignar',
        'Fecha': t.created_at.strftime('%Y-%m-%d %H:%M:%S')
    } for t in tickets]
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    # Use Pandas to write Excel but access the workbook to add filters and logo
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write data starting from row 4 to leave space for logo
        df.to_excel(writer, index=False, sheet_name='Tickets', startrow=3)
        worksheet = writer.sheets['Tickets']
        workbook = writer.book
        
        # Add logo image
        from openpyxl.drawing.image import Image as XLImage
        import os
        from flask import current_app
        logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 50  # Adjust size
            img.width = 50
            worksheet.add_image(img, 'A1')
        
        # Add title next to logo
        from openpyxl.styles import Font, Alignment
        worksheet['B1'] = 'HELP DESK'
        worksheet['B1'].font = Font(size=20, bold=True)
        worksheet['B1'].alignment = Alignment(vertical='center')
        worksheet['B2'] = 'Reporte de Tickets'
        worksheet['B2'].font = Font(size=14)
        
        # Add auto-filter to data columns (starting from row 4)
        worksheet.auto_filter.ref = f'A4:{worksheet.dimensions.split(":")[1]}'
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    output.seek(0)
    
    return send_file(output, download_name='reporte_tickets.xlsx', as_attachment=True)

@bp.route('/export/csv')
@login_required
def export_csv():
    # Filter based on role
    query = Ticket.query
    if current_user.role == 'tecnico':
        query = query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        query = query.filter_by(created_by_id=current_user.id)
        
    tickets = query.all()
    data = [{
        'ID': t.id,
        'Título': t.title,
        'Estado': t.status,
        'Prioridad': t.priority,
        'Creado Por': t.created_by.username,
        'Asignado A': t.assigned_to.username if t.assigned_to else 'Sin asignar',
        'Fecha': t.created_at.strftime('%Y-%m-%d %H:%M:%S')
    } for t in tickets]
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    
    return send_file(output, download_name='reporte_tickets.csv', as_attachment=True, mimetype='text/csv')

@bp.route('/export/pdf')
@login_required
def export_pdf():
    # Filter based on role
    query = Ticket.query
    if current_user.role == 'tecnico':
        query = query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'usuario':
        query = query.filter_by(created_by_id=current_user.id)
        
    tickets = query.all()
    
    pdf = FPDF()
    pdf.add_page()
    
    # Add logo
    import os
    from flask import current_app
    logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
    if os.path.exists(logo_path):
        # Logo on the left
        pdf.image(logo_path, x=10, y=8, w=20)
    
    # Title next to logo
    pdf.set_font("Arial", 'B', 20)
    pdf.set_xy(35, 10)
    pdf.cell(0, 10, txt="HELP DESK", ln=1)
    pdf.set_font("Arial", 'B', 14)
    pdf.set_x(35)
    pdf.cell(0, 5, txt="Reporte de Tickets", ln=1)
    pdf.ln(10)
    
    # Table Header
    pdf.set_font("Arial", 'B', 10)
    # Column widths: ID, Title, Status, Priority, CreatedBy, AssignedTo, Date
    # Total A4 width ~190mm margins included
    col_widths = [10, 40, 25, 20, 25, 25, 45]
    headers = ['ID', 'Título', 'Estado', 'Prioridad', 'Creado Por', 'Asignado A', 'Fecha']
    
    for i in range(len(headers)):
        pdf.cell(col_widths[i], 10, headers[i], border=1, align='C')
    pdf.ln()
    
    # Table Body
    pdf.set_font("Arial", size=10)
    for t in tickets:
        assigned = t.assigned_to.username if t.assigned_to else 'Sin asignar'
        date_str = t.created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        data = [str(t.id), t.title, t.status, t.priority, t.created_by.username, assigned, date_str]
        
        # Check for max height needed (simple approach, just cut off if too long or let it flow? 
        # FPDF cell doesn't wrap automatically without MultiCell, but for this simple report Cell is safer for alignment)
        # We will use Cell and truncate if necessary or just let it be. 
        # For better table we would need a more complex logic, but this matches the request for "table look".
        
        for i in range(len(data)):
            # Truncate title if too long to avoid messing up table
            text = str(data[i])
            if i == 1 and len(text) > 20:
                text = text[:17] + '...'
            pdf.cell(col_widths[i], 10, text, border=1, align='C')
        pdf.ln()
        
    output = io.BytesIO(pdf.output(dest='S'))
    output.seek(0)
    
    return send_file(output, download_name='reporte_tickets.pdf', as_attachment=True)

@bp.route('/search')
@login_required
def search():
    query = request.args.get('q', '')
    if not query:
        return redirect(url_for('main.index'))
    
    # Search tickets
    tickets = Ticket.query.filter(
        db.or_(
            Ticket.title.ilike(f'%{query}%'),
            Ticket.description.ilike(f'%{query}%'),
            Ticket.id.ilike(f'%{query}%')
        )
    ).all()
    
    # Search users (only for admin/tech)
    users = []
    if current_user.role in ['admin', 'tecnico']:
        users = User.query.filter(
            db.or_(
                User.username.ilike(f'%{query}%'),
                User.email.ilike(f'%{query}%')
            )
        ).all()
        
    return render_template('search_results.html', query=query, tickets=tickets, users=users)

@bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        file = request.files.get('profile_picture')
        if file and file.filename:
            from werkzeug.utils import secure_filename
            import os
            from flask import current_app
            
            # Validate file extension
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            
            if file_ext not in allowed_extensions:
                error('Formato de archivo no permitido. Use PNG, JPG, JPEG o GIF.')
                return redirect(url_for('main.profile'))
            
            # Validate file size (2MB max)
            file.seek(0, os.SEEK_END)
            file_size = file.tell()
            file.seek(0)
            
            if file_size > 2 * 1024 * 1024:  # 2MB
                error('El archivo es demasiado grande. Máximo 2MB.')
                return redirect(url_for('main.profile'))
            
            # Save with unique filename
            filename = f'profile_{current_user.id}_{datetime.now().strftime("%Y%m%d%H%M%S")}.{file_ext}'
            os.makedirs(current_app.config['UPLOAD_FOLDER'], exist_ok=True)
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Update user profile picture
            current_user.profile_picture = filename
            db.session.commit()
            
            success('Foto de perfil actualizada exitosamente!')
        
        return redirect(url_for('main.profile'))
    
    return render_template('profile.html')
